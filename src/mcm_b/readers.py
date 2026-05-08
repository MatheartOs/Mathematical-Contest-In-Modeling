"""Lightweight readers for multi-source heterogeneous office files.

The contest data contains thousands of files, so these helpers are designed for
bounded sampling first. Every extraction function accepts a character budget and
returns status/warnings instead of failing the whole pipeline on one bad file.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
import statistics
import xml.etree.ElementTree as ET
import zipfile
import zlib

from openpyxl import load_workbook


TEXT_EXTENSIONS = {".txt"}
WORD_EXTENSIONS = {".docx"}
SPREADSHEET_EXTENSIONS = {".xlsx"}
PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}


@dataclass
class DocumentRecord:
    """A normalized document representation used by downstream models."""

    doc_id: str
    path: str
    dataset: str
    extension: str
    size_bytes: int
    text: str
    status: str
    metadata: dict[str, object] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


def read_document(path: Path, dataset: str, max_chars: int = 20_000) -> DocumentRecord:
    """Read one document with a conservative per-file character budget."""

    path = Path(path)
    extension = path.suffix.lower()
    warnings: list[str] = []
    metadata: dict[str, object] = {}
    text = ""
    status = "ok"

    try:
        if extension in TEXT_EXTENSIONS:
            text, encoding = _read_txt(path, max_chars)
            metadata["encoding"] = encoding
        elif extension in WORD_EXTENSIONS:
            text = _read_docx(path, max_chars)
        elif extension in SPREADSHEET_EXTENSIONS:
            text, metadata = _read_xlsx(path, max_chars=max_chars)
        elif extension in PDF_EXTENSIONS:
            text, pdf_warnings = _read_pdf(path, max_chars=max_chars)
            warnings.extend(pdf_warnings)
            if not text.strip():
                status = "pdf_no_text"
        elif extension in IMAGE_EXTENSIONS:
            metadata, img_warnings = _read_image_metadata(path)
            warnings.extend(img_warnings)
            status = "image_metadata_only"
        else:
            status = "unsupported_extension"
            warnings.append(f"Unsupported extension: {extension}")
    except Exception as exc:  # noqa: BLE001 - keep batch jobs alive.
        status = "error"
        warnings.append(f"{type(exc).__name__}: {exc}")

    size = path.stat().st_size if path.exists() else 0
    return DocumentRecord(
        doc_id=path.stem,
        path=str(path),
        dataset=dataset,
        extension=extension,
        size_bytes=size,
        text=text[:max_chars],
        status=status,
        metadata=metadata,
        warnings=warnings,
    )


def _read_txt(path: Path, max_chars: int) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding, errors="strict")[:max_chars], encoding
        except UnicodeError:
            continue
    return path.read_text(encoding="gb18030", errors="ignore")[:max_chars], "gb18030-ignore"


def _read_docx(path: Path, max_chars: int) -> str:
    """Extract visible text from docx using the zip/XML format directly."""

    parts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for member in ("word/document.xml",):
            if member not in archive.namelist():
                continue
            root = ET.fromstring(archive.read(member))
            for node in root.iter():
                tag = node.tag.rsplit("}", 1)[-1]
                if tag == "t" and node.text:
                    parts.append(node.text)
                elif tag in {"p", "tr"}:
                    parts.append("\n")
                elif tag == "tab":
                    parts.append("\t")
                if sum(len(part) for part in parts) >= max_chars:
                    break
    return _clean_text("".join(parts))[:max_chars]


def _read_xlsx(
    path: Path,
    max_chars: int,
    max_rows_per_sheet: int = 30,
    max_cols: int = 20,
) -> tuple[str, dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    metadata: dict[str, object] = {
        "sheet_names": workbook.sheetnames,
        "sheets": [],
    }
    chunks: list[str] = []
    for sheet in workbook.worksheets:
        metadata["sheets"].append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }
        )
        chunks.append(f"[sheet] {sheet.title}")
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row or 0, max_rows_per_sheet),
            max_col=min(sheet.max_column or 0, max_cols),
            values_only=True,
        ):
            values = ["" if value is None else str(value) for value in row]
            line = "\t".join(value for value in values if value)
            if line:
                chunks.append(line)
            if sum(len(chunk) for chunk in chunks) >= max_chars:
                break
    return _clean_text("\n".join(chunks))[:max_chars], metadata


def _read_image_metadata(path: Path) -> tuple[dict[str, object], list[str]]:
    warnings: list[str] = []
    try:
        from PIL import Image, ImageStat
    except ImportError:
        return {}, ["Pillow is unavailable; image metadata was not extracted."]

    with Image.open(path) as image:
        metadata: dict[str, object] = {
            "width": image.width,
            "height": image.height,
            "mode": image.mode,
            "format": image.format,
        }
        try:
            stat = ImageStat.Stat(image.convert("L").resize((64, 64)))
            metadata["mean_brightness"] = round(float(stat.mean[0]), 3)
            metadata["brightness_std"] = round(float(stat.stddev[0]), 3)
        except Exception as exc:  # noqa: BLE001 - optional visual statistics.
            warnings.append(f"Image statistics failed: {exc}")
    return metadata, warnings


def _read_pdf(path: Path, max_chars: int) -> tuple[str, list[str]]:
    """Extract text from common text PDFs.

    If pypdf is installed, use it. Otherwise fall back to a minimal ToUnicode
    extractor that works for many Office/WPS-generated PDFs in this contest.
    Scanned/image PDFs will intentionally return little or no text.
    """

    warnings: list[str] = []
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        return _clean_text(text)[:max_chars], warnings
    except ImportError:
        warnings.append("pypdf not installed; using minimal built-in PDF extractor.")
    except Exception as exc:  # noqa: BLE001
        warnings.append(f"pypdf extraction failed: {exc}; using built-in fallback.")

    try:
        return _extract_pdf_text_minimal(path, max_chars=max_chars), warnings
    except Exception as exc:  # noqa: BLE001
        warnings.append(f"built-in PDF extraction failed: {type(exc).__name__}: {exc}")
        return "", warnings


def _extract_pdf_text_minimal(path: Path, max_chars: int) -> str:
    data = path.read_bytes()
    objects = _parse_pdf_objects(data)
    font_cmaps = _pdf_font_cmaps(objects)
    page_numbers = [
        number
        for number, body in sorted(objects.items())
        if re.search(rb"/Type\s*/Page\b", body)
    ]

    text_parts: list[str] = []
    for page_number in page_numbers:
        text_parts.append(_extract_pdf_page_text(objects, font_cmaps, page_number))
        if sum(len(part) for part in text_parts) >= max_chars:
            break
    return _clean_text("\n".join(text_parts))[:max_chars]


def _parse_pdf_objects(data: bytes) -> dict[int, bytes]:
    objects: dict[int, bytes] = {}
    pattern = re.compile(rb"(\d+)\s+(\d+)\s+obj\b(.*?)\bendobj", re.S)
    for match in pattern.finditer(data):
        objects[int(match.group(1))] = match.group(3)
    return objects


def _pdf_stream(body: bytes) -> bytes:
    if b"stream" not in body:
        return b""
    header, rest = body.split(b"stream", 1)
    raw = rest.split(b"endstream", 1)[0]
    candidates = [
        raw.strip(b"\r\n"),
        raw.lstrip(b"\r\n"),
        raw.lstrip(b"\r\n").rstrip(b"\n"),
        raw.lstrip(b"\r\n") + b"\r",
        raw.lstrip(b"\r\n") + b"\r\n",
    ]
    if b"/FlateDecode" not in header:
        return raw.strip(b"\r\n")
    last_error: Exception | None = None
    for candidate in candidates:
        try:
            return zlib.decompress(candidate)
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    if last_error:
        raise last_error
    return b""


def _pdf_font_cmaps(objects: dict[int, bytes]) -> dict[int, dict[str, str]]:
    cmaps: dict[int, dict[str, str]] = {}
    for object_number, body in objects.items():
        match = re.search(rb"/ToUnicode\s+(\d+)\s+0\s+R", body)
        if not match:
            continue
        cmap_object = int(match.group(1))
        cmaps[object_number] = _parse_pdf_cmap(_pdf_stream(objects[cmap_object]))
    return cmaps


def _parse_pdf_cmap(stream: bytes) -> dict[str, str]:
    cmap: dict[str, str] = {}
    text = stream.decode("latin-1", errors="ignore")
    for block in re.finditer(r"beginbfchar(.*?)endbfchar", text, flags=re.S):
        for source, target in re.findall(
            r"<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>",
            block.group(1),
        ):
            cmap[source.upper()] = _decode_pdf_unicode_hex(target)

    for block in re.finditer(r"beginbfrange(.*?)endbfrange", text, flags=re.S):
        for source, end, target in re.findall(
            r"<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>",
            block.group(1),
        ):
            start_code = int(source, 16)
            end_code = int(end, 16)
            target_code = int(target, 16)
            width = len(source)
            for code in range(start_code, end_code + 1):
                mapped = target_code + code - start_code
                cmap[f"{code:0{width}X}"] = _decode_pdf_unicode_hex(f"{mapped:04X}")
    return cmap


def _decode_pdf_unicode_hex(value: str) -> str:
    try:
        return bytes.fromhex(value).decode("utf-16-be")
    except Exception:  # noqa: BLE001
        return chr(int(value, 16))


def _extract_pdf_page_text(
    objects: dict[int, bytes],
    font_cmaps: dict[int, dict[str, str]],
    page_number: int,
) -> str:
    page = objects[page_number]
    content_refs = _content_refs(page)
    font_refs = {
        name.decode("ascii", errors="ignore"): int(number)
        for name, number in re.findall(rb"/(FT\d+|F\d+)\s+(\d+)\s+0\s+R", page)
    }
    current_font: str | None = None
    chunks: list[str] = []
    token_pattern = re.compile(
        rb"/(FT\d+|F\d+)\s+[-0-9.]+\s+Tf"
        rb"|<([0-9A-Fa-f]+)>\s*Tj"
        rb"|\[(.*?)\]\s*TJ"
        rb"|\((.*?)\)\s*Tj"
        rb"|\bET\b",
        re.S,
    )

    for content_ref in content_refs:
        content = _pdf_stream(objects[content_ref])
        for match in token_pattern.finditer(content):
            if match.group(1):
                current_font = match.group(1).decode("ascii", errors="ignore")
            elif match.group(2):
                chunks.append(
                    _decode_pdf_hex_text(
                        match.group(2).decode("ascii"),
                        font_cmaps.get(font_refs.get(current_font or "", -1), {}),
                    )
                )
            elif match.group(3):
                cmap = font_cmaps.get(font_refs.get(current_font or "", -1), {})
                for hex_text in re.findall(rb"<([0-9A-Fa-f]+)>", match.group(3)):
                    chunks.append(_decode_pdf_hex_text(hex_text.decode("ascii"), cmap))
            elif match.group(4):
                chunks.append(_decode_pdf_literal(match.group(4)))
            else:
                chunks.append("\n")
    return "".join(chunks)


def _content_refs(page: bytes) -> list[int]:
    list_match = re.search(rb"/Contents\s*\[(.*?)\]", page, flags=re.S)
    if list_match:
        return [int(value) for value in re.findall(rb"(\d+)\s+0\s+R", list_match.group(1))]
    single_match = re.search(rb"/Contents\s+(\d+)\s+0\s+R", page)
    return [int(single_match.group(1))] if single_match else []


def _decode_pdf_hex_text(value: str, cmap: dict[str, str]) -> str:
    value = value.upper()
    output: list[str] = []
    i = 0
    while i < len(value):
        four = value[i : i + 4]
        two = value[i : i + 2]
        if four in cmap:
            output.append(cmap[four])
            i += 4
        elif two in cmap:
            output.append(cmap[two])
            i += 2
        else:
            try:
                output.append(bytes.fromhex(four).decode("utf-16-be"))
            except Exception:  # noqa: BLE001
                output.append("")
            i += 4
    return "".join(output)


def _decode_pdf_literal(value: bytes) -> str:
    value = value.replace(rb"\(", b"(").replace(rb"\)", b")").replace(rb"\\", b"\\")
    for encoding in ("utf-8", "gb18030", "latin-1"):
        try:
            return value.decode(encoding)
        except UnicodeError:
            continue
    return value.decode("latin-1", errors="ignore")


def _clean_text(text: str) -> str:
    text = text.replace("\x00", "")
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r" *\n+ *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def text_quality_score(text: str) -> float:
    """Heuristic score for whether extracted text is useful enough to model."""

    if not text:
        return 0.0
    visible = sum(1 for char in text if char.isprintable() and not char.isspace())
    chinese = sum(1 for char in text if "\u4e00" <= char <= "\u9fff")
    alpha_numeric = sum(1 for char in text if char.isalpha() or char.isdigit())
    density = visible / max(len(text), 1)
    language_signal = (chinese + alpha_numeric) / max(visible, 1)
    return round(min(1.0, density * language_signal * 1.2), 4)


def file_size_summary(records: list[DocumentRecord]) -> dict[str, float]:
    """Return small descriptive statistics for sampled records."""

    sizes = [record.size_bytes for record in records]
    if not sizes:
        return {"count": 0}
    return {
        "count": float(len(sizes)),
        "min": float(min(sizes)),
        "median": float(statistics.median(sizes)),
        "max": float(max(sizes)),
    }
