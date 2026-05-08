"""Document cleaning pipeline for multi-source heterogeneous office files.

This module follows the workflow in
``docs/多源异构文件数据清洗流程说明文档.md``.  The primary output is a
file-level ``document_index.csv`` plus block-level ``document_blocks.jsonl``.
OCR is represented as an explicit pending/manual-check path when no OCR engine
is available in the local environment.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import datetime
from functools import lru_cache
import json
import os
from pathlib import Path
import re
import statistics
import xml.etree.ElementTree as ET
import zipfile
import zlib

import pandas as pd
from openpyxl import load_workbook

from .paths import DATASET1_DIR, DATASET2_DIR, DATASET3_XLSX, DATASET4_XLSX


TEXT_EXTENSIONS = {".txt"}
WORD_EXTENSIONS = {".docx"}
EXCEL_EXTENSIONS = {".xlsx", ".xls", ".csv"}
PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}

PROJECT_MODEL_ROOT = Path(__file__).resolve().parents[2] / "models" / "paddleocr"
ASCII_MODEL_ROOT = Path(r"C:\mcm_paddleocr_models")

BUSINESS_DICTIONARY: dict[str, tuple[str, ...]] = {
    "notice": ("通知", "公告", "公示", "安排", "要求"),
    "meeting": ("会议", "参会", "议程", "纪要", "签到"),
    "project": ("项目", "申报", "立项", "验收", "结题", "成果"),
    "money": ("经费", "预算", "报销", "拨款", "补贴", "金额", "资金"),
    "contract": ("合同", "协议", "甲方", "乙方", "责任书", "条款"),
    "personnel": ("招聘", "考核", "任免", "培训", "名单"),
    "deadline": ("截止", "限期", "前", "之前", "逾期", "报送", "提交", "上报"),
    "urgent": ("紧急", "立即", "尽快", "马上", "务必", "优先"),
}

BUSINESS_SYNONYMS: dict[str, tuple[str, ...]] = {
    "资金": ("经费", "预算", "拨款", "补贴", "报销"),
    "会议": ("会议安排", "会议通知", "参会"),
    "项目": ("申报", "立项", "验收", "结题"),
    "合同": ("协议", "责任书"),
    "通知": ("公告", "公示"),
}

ORG_SUFFIXES = ("学院", "部", "处", "办公室", "中心", "委员会", "公司", "单位", "小组", "项目组", "局", "厅")


@dataclass
class FileManifestRow:
    file_id: str
    original_id: str
    dataset_id: str
    file_name: str
    file_path: str
    file_ext: str
    file_size_kb: float
    modified_time: str
    is_readable: int
    initial_type: str
    parse_status: str = "pending"


@dataclass
class DocumentBlock:
    file_id: str
    block_id: str
    page: int | None
    block_type: str
    text: str
    bbox: list[float] | None
    source: str
    confidence: float
    reading_order: int


@dataclass
class CleanedDocument:
    file_id: str
    original_id: str
    dataset_id: str
    file_name: str
    file_path: str
    file_type: str
    file_size_kb: float
    parse_method: str
    parse_success: int
    error_message: str
    title: str
    full_text: str
    clean_text: str
    text_length: int
    page_count: int
    paragraph_count: int
    table_count: int
    image_count: int
    heading_count: int
    avg_paragraph_length: float
    max_paragraph_length: int
    has_signature: int
    has_attachment: int
    has_stamp_or_seal: int
    date_list: str
    deadline: str
    has_deadline: int
    amount_list: str
    max_amount: float
    has_money: int
    has_notice: int
    has_meeting: int
    has_project: int
    has_contract: int
    has_personnel: int
    has_urgent: int
    organization_list: str
    contact_list: str
    ocr_used: int
    ocr_confidence: float
    ocr_page_count: int
    parse_quality: float
    missing_rate: float
    need_manual_check: int
    manual_check_reason: str
    keywords_tfidf: str
    keywords_textrank: str
    business_keywords: str
    parse_notes: str = ""


@dataclass
class ParsedContent:
    parse_method: str
    parse_success: int
    error_message: str
    full_text: str
    blocks: list[DocumentBlock]
    page_count: int = 0
    paragraph_count: int = 0
    table_count: int = 0
    image_count: int = 0
    heading_count: int = 0
    ocr_used: int = 0
    ocr_confidence: float = 0.0
    ocr_page_count: int = 0
    parse_notes: str = ""
    layout_score: float = 0.6


def scan_all_inputs() -> list[FileManifestRow]:
    """Scan official datasets and return a unified file manifest."""

    rows: list[FileManifestRow] = []
    rows.extend(_scan_file_directory(DATASET1_DIR, "dataset1", "D1"))
    rows.extend(_scan_file_directory(DATASET2_DIR, "dataset2", "D2"))
    rows.extend(_scan_dataset3())
    rows.extend(_scan_dataset4())
    return rows


def clean_all_documents(
    max_chars: int = 30_000,
    max_file_mb: float = 50.0,
    limits: dict[str, int] | None = None,
) -> tuple[pd.DataFrame, list[DocumentBlock], pd.DataFrame, pd.DataFrame, list[str]]:
    """Run cleaning and return index, blocks, manifest, parse log, errors."""

    manifest = scan_all_inputs()
    if limits:
        manifest = _apply_limits(manifest, limits)

    documents: list[CleanedDocument] = []
    blocks: list[DocumentBlock] = []
    parse_rows: list[dict[str, object]] = []
    errors: list[str] = []
    max_bytes = int(max_file_mb * 1024 * 1024)

    for index, row in enumerate(manifest, start=1):
        content = _parse_manifest_row(row, max_chars=max_chars, max_bytes=max_bytes)
        document = _build_document_index_row(row, content)
        documents.append(document)
        blocks.extend(content.blocks)
        if content.error_message:
            errors.append(f"[{row.file_id}] {content.error_message}")
        parse_rows.append(
            {
                "file_id": row.file_id,
                "file_name": row.file_name,
                "file_type": row.file_ext,
                "parse_method": content.parse_method,
                "parse_success": content.parse_success,
                "text_length": len(content.full_text),
                "table_count": content.table_count,
                "image_count": content.image_count,
                "error_message": content.error_message,
            }
        )
        if index % 500 == 0:
            print(f"[cleaning] parsed {index}/{len(manifest)}")

    document_index = pd.DataFrame(asdict(document) for document in documents)
    manifest_frame = pd.DataFrame(asdict(row) for row in manifest)
    parse_log = pd.DataFrame(parse_rows)
    return document_index, blocks, manifest_frame, parse_log, errors


def write_cleaning_outputs(
    output_dir: Path,
    document_index: pd.DataFrame,
    blocks: list[DocumentBlock],
    manifest: pd.DataFrame,
    parse_log: pd.DataFrame,
    errors: list[str],
) -> None:
    """Write standard cleaning deliverables."""

    processed = output_dir / "processed"
    logs = output_dir / "logs"
    processed.mkdir(parents=True, exist_ok=True)
    logs.mkdir(parents=True, exist_ok=True)

    manifest.to_csv(processed / "file_manifest.csv", index=False, encoding="utf-8-sig")
    document_index.to_csv(processed / "document_index.csv", index=False, encoding="utf-8-sig")
    parse_log.to_csv(logs / "parse_log.csv", index=False, encoding="utf-8-sig")
    _write_blocks_jsonl(processed / "document_blocks.jsonl", blocks)
    _write_business_dictionary(processed / "business_dictionary.json")
    _write_manual_check_list(processed / "manual_check_list.csv", document_index)
    (logs / "error_log.txt").write_text("\n".join(errors), encoding="utf-8")
    _write_ocr_log(logs / "ocr_log.csv", document_index)


def _scan_file_directory(directory: Path, dataset_id: str, prefix: str) -> list[FileManifestRow]:
    rows: list[FileManifestRow] = []
    for index, path in enumerate(sorted(directory.iterdir()), start=1):
        if not path.is_file():
            continue
        stat = path.stat()
        rows.append(
            FileManifestRow(
                file_id=f"{prefix}_{index:04d}",
                original_id=path.stem,
                dataset_id=dataset_id,
                file_name=path.name,
                file_path=str(path),
                file_ext=path.suffix.lower(),
                file_size_kb=round(stat.st_size / 1024, 3),
                modified_time=datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                is_readable=int(_is_readable(path)),
                initial_type=_initial_type(path.suffix.lower()),
            )
        )
    return rows


def _scan_dataset3() -> list[FileManifestRow]:
    frame = pd.read_excel(DATASET3_XLSX)
    rows: list[FileManifestRow] = []
    size_kb = round(DATASET3_XLSX.stat().st_size / 1024, 3)
    modified = datetime.fromtimestamp(DATASET3_XLSX.stat().st_mtime).isoformat(timespec="seconds")
    for index, row in frame.iterrows():
        original_id = str(row.get("文件编号", f"N{index + 1:05d}"))
        rows.append(
            FileManifestRow(
                file_id=f"D3_{index + 1:04d}",
                original_id=original_id,
                dataset_id="dataset3",
                file_name=f"{original_id}.xlsx-row",
                file_path=f"{DATASET3_XLSX}#{original_id}",
                file_ext=".xlsx-row",
                file_size_kb=size_kb,
                modified_time=modified,
                is_readable=1,
                initial_type="semi_structured_row",
            )
        )
    return rows


def _scan_dataset4() -> list[FileManifestRow]:
    stat = DATASET4_XLSX.stat()
    return [
        FileManifestRow(
            file_id="D4_0001",
            original_id=DATASET4_XLSX.stem,
            dataset_id="dataset4",
            file_name=DATASET4_XLSX.name,
            file_path=str(DATASET4_XLSX),
            file_ext=".xlsx",
            file_size_kb=round(stat.st_size / 1024, 3),
            modified_time=datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            is_readable=1,
            initial_type="excel",
        )
    ]


def _apply_limits(rows: list[FileManifestRow], limits: dict[str, int]) -> list[FileManifestRow]:
    limited: list[FileManifestRow] = []
    counts: dict[str, int] = {}
    for row in rows:
        limit = limits.get(row.dataset_id, 0)
        counts[row.dataset_id] = counts.get(row.dataset_id, 0) + 1
        if limit and counts[row.dataset_id] > limit:
            continue
        limited.append(row)
    return limited


def _parse_manifest_row(row: FileManifestRow, max_chars: int, max_bytes: int) -> ParsedContent:
    if row.dataset_id == "dataset3":
        return _parse_dataset3_row(row, max_chars=max_chars)
    path = Path(row.file_path)
    if not path.exists():
        return ParsedContent("missing_file", 0, "File does not exist.", "", [])
    if path.stat().st_size > max_bytes:
        return ParsedContent(
            "metadata_only",
            0,
            f"File exceeds max size budget: {round(max_bytes / 1024 / 1024, 1)} MB.",
            "",
            [],
            parse_notes="large_file_skipped",
            layout_score=0.2,
        )
    ext = row.file_ext
    try:
        if ext in TEXT_EXTENSIONS:
            return _parse_txt(row, max_chars=max_chars)
        if ext in WORD_EXTENSIONS:
            return _parse_docx(row, max_chars=max_chars)
        if ext in EXCEL_EXTENSIONS:
            return _parse_excel(row, max_chars=max_chars)
        if ext in PDF_EXTENSIONS:
            return _parse_pdf(row, max_chars=max_chars)
        if ext in IMAGE_EXTENSIONS:
            return _parse_image(row)
        return ParsedContent("unsupported", 0, f"Unsupported extension: {ext}", "", [])
    except Exception as exc:  # noqa: BLE001 - record and continue batch cleaning.
        return ParsedContent("parse_failed", 0, f"{type(exc).__name__}: {exc}", "", [])


def _parse_txt(row: FileManifestRow, max_chars: int) -> ParsedContent:
    path = Path(row.file_path)
    text, encoding = _read_txt(path)
    clean = normalize_text(text)[:max_chars]
    paragraphs = _split_paragraphs(clean)
    blocks = _paragraph_blocks(row.file_id, paragraphs, "txt")
    if _is_image_sidecar_text(clean):
        return ParsedContent(
            parse_method=f"image_sidecar_txt:{encoding}",
            parse_success=0,
            error_message="Image sidecar metadata text; original image/OCR content required for modeling.",
            full_text=clean,
            blocks=blocks,
            paragraph_count=len(paragraphs),
            heading_count=0,
            parse_notes="image_sidecar_metadata",
            layout_score=0.2,
        )
    return ParsedContent(
        parse_method=f"txt_parse:{encoding}",
        parse_success=1,
        error_message="",
        full_text=clean,
        blocks=blocks,
        paragraph_count=len(paragraphs),
        heading_count=sum(_is_heading(paragraph) for paragraph in paragraphs),
        parse_notes="",
        layout_score=0.7,
    )


def _parse_docx(row: FileManifestRow, max_chars: int) -> ParsedContent:
    path = Path(row.file_path)
    paragraphs: list[str] = []
    tables: list[list[list[str]]] = []
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        if "word/document.xml" not in names:
            raise ValueError("word/document.xml missing")
        root = ET.fromstring(archive.read("word/document.xml"))
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        for paragraph in root.findall(".//w:p", ns):
            texts = [node.text for node in paragraph.findall(".//w:t", ns) if node.text]
            line = normalize_text("".join(texts))
            if line:
                paragraphs.append(line)
        for table in root.findall(".//w:tbl", ns):
            table_rows: list[list[str]] = []
            for tr in table.findall(".//w:tr", ns):
                cells: list[str] = []
                for tc in tr.findall("./w:tc", ns):
                    cell_text = normalize_text("".join(node.text or "" for node in tc.findall(".//w:t", ns)))
                    cells.append(cell_text)
                if any(cells):
                    table_rows.append(cells)
            if table_rows:
                tables.append(table_rows)
    blocks: list[DocumentBlock] = []
    order = 1
    for block in _paragraph_blocks(row.file_id, paragraphs, "docx", start_order=order):
        blocks.append(block)
        order = block.reading_order + 1
    table_texts = [_summarize_table(table) for table in tables]
    for table_index, table_text in enumerate(table_texts, start=1):
        blocks.append(
            DocumentBlock(row.file_id, f"{row.file_id}_T{table_index:02d}", None, "table", table_text, None, "docx_table", 1.0, order)
        )
        order += 1
    full_text = normalize_text("\n".join([*paragraphs, *table_texts]))[:max_chars]
    return ParsedContent(
        "docx_parse",
        1,
        "",
        full_text,
        blocks,
        paragraph_count=len(paragraphs),
        table_count=len(tables),
        image_count=_docx_image_count(path),
        heading_count=sum(_is_heading(paragraph) for paragraph in paragraphs),
        layout_score=0.9 if tables else 0.8,
    )


def _parse_excel(row: FileManifestRow, max_chars: int) -> ParsedContent:
    path = Path(row.file_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    blocks: list[DocumentBlock] = []
    texts: list[str] = []
    order = 1
    table_count = 0
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 0, 60), max_col=min(sheet.max_column or 0, 20), values_only=True))
        if not rows:
            continue
        headers = _infer_headers(rows)
        sheet_text = _excel_sheet_text(sheet.title, sheet.max_row or 0, sheet.max_column or 0, headers, rows)
        texts.append(sheet_text)
        table_count += 1
        blocks.append(DocumentBlock(row.file_id, f"{row.file_id}_S{table_count:02d}", None, "table", sheet_text, None, "excel", 1.0, order))
        order += 1
        if sum(len(text) for text in texts) >= max_chars:
            break
    full_text = normalize_text("\n".join(texts))[:max_chars]
    return ParsedContent(
        "excel_parse",
        1,
        "",
        full_text,
        blocks,
        table_count=table_count,
        layout_score=0.9,
    )


def _parse_pdf(row: FileManifestRow, max_chars: int) -> ParsedContent:
    path = Path(row.file_path)
    page_count = _pdf_page_count(path)
    text = _extract_pdf_text_minimal(path, max_chars=max_chars)
    text_per_page = len(text) / max(page_count, 1)
    if text_per_page >= 100:
        method = "text_pdf"
        note = ""
        success = 1
        layout_score = 0.65
    elif text_per_page >= 20:
        method = "mixed_pdf_text_only"
        note = "mixed_pdf_needs_low_text_page_ocr"
        success = 1
        layout_score = 0.55
    elif text:
        method = "scanned_pdf_text_low"
        note = "scanned_pdf_needs_ocr"
        success = 1
        layout_score = 0.4
    else:
        method = "scanned_pdf_ocr_pending"
        note = "no_text_extracted_ocr_required"
        success = 0
        layout_score = 0.2
    blocks = []
    if text:
        blocks.append(DocumentBlock(row.file_id, f"{row.file_id}_P001_B001", 1, "paragraph", text, None, "pdf_text", 1.0, 1))
    return ParsedContent(
        method,
        success,
        "" if success else "PDF text extraction too short; OCR required.",
        normalize_text(text),
        blocks,
        page_count=page_count,
        paragraph_count=len(_split_paragraphs(text)),
        ocr_used=0,
        ocr_confidence=0.0,
        ocr_page_count=0,
        parse_notes=note,
        layout_score=layout_score,
    )


def _parse_image(row: FileManifestRow) -> ParsedContent:
    path = Path(row.file_path)
    notes = []
    width = height = 0
    quality = "unknown"
    try:
        from PIL import Image, ImageStat

        with Image.open(path) as image:
            width, height = image.width, image.height
            stat = ImageStat.Stat(image.convert("L").resize((64, 64)))
            blur_proxy = float(stat.stddev[0])
            quality = "high" if width * height >= 800_000 and blur_proxy >= 35 else "medium" if width * height >= 250_000 else "low"
    except Exception as exc:  # noqa: BLE001
        notes.append(f"image_quality_failed:{exc}")
    metadata_text = f"image_width={width}; image_height={height}; image_quality={quality}"
    try:
        ocr_text, ocr_blocks, avg_confidence = _run_paddleocr_image(path, row.file_id)
    except Exception as exc:  # noqa: BLE001 - OCR errors must be logged per file.
        block = DocumentBlock(row.file_id, f"{row.file_id}_IMG001", None, "image", metadata_text, None, "image_metadata", 1.0, 1)
        return ParsedContent(
            "image_ocr_failed",
            0,
            f"PaddleOCR failed: {type(exc).__name__}: {exc}",
            "",
            [block],
            image_count=1,
            ocr_used=1,
            ocr_confidence=0.0,
            ocr_page_count=1,
            parse_notes=";".join(notes + [f"image_quality={quality}"]),
            layout_score=0.25,
        )
    blocks = [DocumentBlock(row.file_id, f"{row.file_id}_IMG_META", None, "image", metadata_text, None, "image_metadata", 1.0, 1)]
    blocks.extend(ocr_blocks)
    if not ocr_text.strip():
        return ParsedContent(
            "image_ocr_empty",
            0,
            "PaddleOCR returned no text; manual check required.",
            "",
            blocks,
            image_count=1,
            ocr_used=1,
            ocr_confidence=avg_confidence,
            ocr_page_count=1,
            parse_notes=";".join(notes + [f"image_quality={quality}"]),
            layout_score=0.35,
        )
    return ParsedContent(
        "image_paddleocr",
        1,
        "",
        normalize_text(ocr_text),
        blocks,
        image_count=1,
        paragraph_count=len(_split_paragraphs(ocr_text)),
        ocr_used=1,
        ocr_confidence=avg_confidence,
        ocr_page_count=1,
        parse_notes=";".join(notes + [f"image_quality={quality}"]),
        layout_score=0.55,
    )


def _parse_dataset3_row(row: FileManifestRow, max_chars: int) -> ParsedContent:
    doc_id = row.original_id
    frame = _dataset3_frame()
    record = frame[frame["文件编号"].astype(str) == doc_id]
    if record.empty:
        return ParsedContent("dataset3_missing_row", 0, "Dataset3 row not found.", "", [])
    item = record.iloc[0]
    text = "" if pd.isna(item.get("正文片段", "")) else str(item.get("正文片段", ""))
    time_info = "" if pd.isna(item.get("时间信息", "")) else str(item.get("时间信息", ""))
    text = normalize_text(f"{text}\n时间信息：{time_info}")[:max_chars]
    block = DocumentBlock(row.file_id, f"{row.file_id}_B001", None, "paragraph", text, None, "dataset3_row", 1.0, 1)
    return ParsedContent("dataset3_row_parse", 1, "", text, [block], paragraph_count=1, layout_score=0.6)


@lru_cache(maxsize=1)
def _dataset3_frame() -> pd.DataFrame:
    return pd.read_excel(DATASET3_XLSX)


def _build_document_index_row(row: FileManifestRow, content: ParsedContent) -> CleanedDocument:
    clean_text = normalize_text(content.full_text)
    paragraphs = _split_paragraphs(clean_text)
    title = _extract_title(row.file_name, paragraphs)
    dates = extract_dates(clean_text)
    deadline = extract_deadline(clean_text, dates)
    amounts = extract_amounts(clean_text)
    flags, matched_terms = business_flags(clean_text)
    orgs = extract_organizations(clean_text)
    contacts = extract_contacts(clean_text)
    keywords = extract_keywords(clean_text, top_n=12)
    business_keywords = sorted({term for terms in matched_terms.values() for term in terms})
    quality, missing_rate, manual, reason = quality_score(row, content, clean_text, title, dates, amounts, flags)
    paragraph_lengths = [len(paragraph) for paragraph in paragraphs]
    return CleanedDocument(
        file_id=row.file_id,
        original_id=row.original_id,
        dataset_id=row.dataset_id,
        file_name=row.file_name,
        file_path=row.file_path,
        file_type=row.file_ext,
        file_size_kb=row.file_size_kb,
        parse_method=content.parse_method,
        parse_success=content.parse_success,
        error_message=content.error_message,
        title=title,
        full_text=content.full_text,
        clean_text=clean_text,
        text_length=len(clean_text),
        page_count=content.page_count,
        paragraph_count=content.paragraph_count or len(paragraphs),
        table_count=content.table_count,
        image_count=content.image_count,
        heading_count=content.heading_count,
        avg_paragraph_length=round(statistics.mean(paragraph_lengths), 3) if paragraph_lengths else 0.0,
        max_paragraph_length=max(paragraph_lengths) if paragraph_lengths else 0,
        has_signature=int(bool(re.search(r"(特此|此致|联系人|联系电话|年\d{1,2}月\d{1,2}日|盖章)", clean_text))),
        has_attachment=int("附件" in clean_text),
        has_stamp_or_seal=int(any(term in clean_text for term in ("盖章", "印章", "签章", "公章"))),
        date_list=json.dumps(dates, ensure_ascii=False),
        deadline=deadline,
        has_deadline=flags["has_deadline"],
        amount_list=json.dumps(amounts, ensure_ascii=False),
        max_amount=max(amounts) if amounts else 0.0,
        has_money=flags["has_money"],
        has_notice=flags["has_notice"],
        has_meeting=flags["has_meeting"],
        has_project=flags["has_project"],
        has_contract=flags["has_contract"],
        has_personnel=flags["has_personnel"],
        has_urgent=flags["has_urgent"],
        organization_list=json.dumps(orgs, ensure_ascii=False),
        contact_list=json.dumps(contacts, ensure_ascii=False),
        ocr_used=content.ocr_used,
        ocr_confidence=content.ocr_confidence,
        ocr_page_count=content.ocr_page_count,
        parse_quality=quality,
        missing_rate=missing_rate,
        need_manual_check=manual,
        manual_check_reason=reason,
        keywords_tfidf=json.dumps(keywords, ensure_ascii=False),
        keywords_textrank=json.dumps(keywords, ensure_ascii=False),
        business_keywords=json.dumps(business_keywords, ensure_ascii=False),
        parse_notes=content.parse_notes,
    )


def normalize_text(text: str) -> str:
    text = _fullwidth_to_halfwidth(text)
    for source, targets in BUSINESS_SYNONYMS.items():
        for target in targets:
            text = text.replace(target, source)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n+ *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_dates(text: str) -> list[str]:
    found: list[str] = []
    patterns = [
        r"((?:19|20)\d{2})年\s*(\d{1,2})月\s*(\d{1,2})日?",
        r"((?:19|20)\d{2})[./-](\d{1,2})[./-](\d{1,2})",
    ]
    for pattern in patterns:
        for year, month, day in re.findall(pattern, text):
            found.append(f"{int(year):04d}-{int(month):02d}-{int(day):02d}")
    for month, day in re.findall(r"(?<!\d)(\d{1,2})月\s*(\d{1,2})日?(?:前|之前)?", text):
        found.append(f"2026-{int(month):02d}-{int(day):02d}")
    return sorted(set(found))


def extract_deadline(text: str, dates: list[str]) -> str:
    if not dates:
        return ""
    for date in dates:
        variants = [date, date.replace("-", "年", 1).replace("-", "月", 1) + "日"]
        for variant in variants:
            index = text.find(variant)
            if index >= 0:
                context = text[max(0, index - 20) : index + 30]
                if any(term in context for term in BUSINESS_DICTIONARY["deadline"]):
                    return date
    return ""


def extract_amounts(text: str) -> list[float]:
    amounts: list[float] = []
    for value, unit in re.findall(r"(\d+(?:,\d{3})*(?:\.\d+)?)\s*(万元|千元|元)", text):
        number = float(value.replace(",", ""))
        if unit == "万元":
            number *= 10_000
        elif unit == "千元":
            number *= 1_000
        amounts.append(round(number, 2))
    return amounts


def business_flags(text: str) -> tuple[dict[str, int], dict[str, list[str]]]:
    flags: dict[str, int] = {}
    matched: dict[str, list[str]] = {}
    for key, terms in BUSINESS_DICTIONARY.items():
        hits = sorted({term for term in terms if term in text})
        output_key = f"has_{key}"
        flags[output_key] = int(bool(hits))
        matched[key] = hits
    return flags, matched


def extract_organizations(text: str) -> list[str]:
    orgs: set[str] = set()
    for suffix in ORG_SUFFIXES:
        for match in re.findall(rf"[\u4e00-\u9fffA-Za-z0-9]{{2,30}}{suffix}", text):
            orgs.add(match[-40:])
    return sorted(orgs)[:20]


def extract_contacts(text: str) -> list[str]:
    contacts = set(re.findall(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", text))
    contacts.update(re.findall(r"(?<!\d)(?:1[3-9]\d{9}|\d{3,4}-?\d{7,8})(?!\d)", text))
    return sorted(contacts)[:20]


def extract_keywords(text: str, top_n: int) -> list[str]:
    candidates: dict[str, int] = {}
    for terms in BUSINESS_DICTIONARY.values():
        for term in terms:
            count = text.count(term)
            if count:
                candidates[term] = candidates.get(term, 0) + count
    for word in re.findall(r"[\u4e00-\u9fff]{2,6}", text):
        if len(word) >= 2:
            candidates[word] = candidates.get(word, 0) + 1
    return [word for word, _ in sorted(candidates.items(), key=lambda item: (-item[1], item[0]))[:top_n]]


def quality_score(
    row: FileManifestRow,
    content: ParsedContent,
    clean_text: str,
    title: str,
    dates: list[str],
    amounts: list[float],
    flags: dict[str, int],
) -> tuple[float, float, int, str]:
    q_text = min(len(clean_text) / 800, 1.0) if clean_text else 0.0
    q_ocr = 1.0 if not content.ocr_used and content.parse_success else content.ocr_confidence
    q_layout = content.layout_score
    key_fields = [bool(title), bool(clean_text), bool(dates), bool(amounts), any(flags.values())]
    missing = key_fields.count(False)
    q_complete = 1 - missing / len(key_fields)
    q_format = 1.0 if content.parse_success else 0.0
    score = round(0.30 * q_text + 0.20 * q_ocr + 0.20 * q_layout + 0.20 * q_complete + 0.10 * q_format, 4)
    reasons: list[str] = []
    if score < 0.5:
        reasons.append("parse_quality_below_0.5")
    elif score < 0.7:
        reasons.append("parse_quality_between_0.5_and_0.7")
    if len(clean_text) < 80:
        reasons.append("text_too_short")
    if row.file_ext in IMAGE_EXTENSIONS or "ocr" in content.parse_method:
        reasons.append("ocr_required_or_pending")
    if content.error_message:
        reasons.append(content.error_message)
    return score, round(missing / len(key_fields), 4), int(bool(reasons)), "; ".join(reasons)


def _write_blocks_jsonl(path: Path, blocks: list[DocumentBlock]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for block in blocks:
            handle.write(json.dumps(asdict(block), ensure_ascii=False) + "\n")


def _write_business_dictionary(path: Path) -> None:
    payload = {key: list(values) for key, values in BUSINESS_DICTIONARY.items()}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_manual_check_list(path: Path, document_index: pd.DataFrame) -> None:
    columns = ["file_id", "dataset_id", "file_name", "file_type", "parse_quality", "manual_check_reason", "file_path"]
    document_index[document_index["need_manual_check"] == 1][columns].to_csv(path, index=False, encoding="utf-8-sig")


def _write_ocr_log(path: Path, document_index: pd.DataFrame) -> None:
    rows: list[dict[str, object]] = []
    ocr_frame = document_index[document_index["ocr_used"] == 1].copy()
    for _, row in ocr_frame.iterrows():
        rows.append(
            {
                "file_id": row["file_id"],
                "page": 1 if int(row.get("ocr_page_count", 0) or 0) else "",
                "ocr_used": row["ocr_used"],
                "ocr_confidence": row["ocr_confidence"],
                "image_quality": _parse_note_value(str(row.get("parse_notes", "")), "image_quality"),
                "ocr_text_length": row["text_length"],
                "ocr_error": row["error_message"] if int(row.get("parse_success", 0)) == 0 else "",
            }
        )
    pd.DataFrame(
        rows,
        columns=["file_id", "page", "ocr_used", "ocr_confidence", "image_quality", "ocr_text_length", "ocr_error"],
    ).to_csv(path, index=False, encoding="utf-8-sig")


def _parse_note_value(notes: str, key: str) -> str:
    for part in notes.split(";"):
        if part.startswith(f"{key}="):
            return part.split("=", 1)[1]
    return ""


@lru_cache(maxsize=1)
def _paddleocr_engine():
    det_dir, rec_dir = _resolve_paddleocr_model_dirs()
    device = _resolve_paddleocr_device()
    from paddleocr import PaddleOCR

    return PaddleOCR(
        text_detection_model_dir=str(det_dir),
        text_recognition_model_dir=str(rec_dir),
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
        device=device,
        enable_mkldnn=False,
    )


def _resolve_paddleocr_device() -> str:
    env_device = os.environ.get("PADDLEOCR_DEVICE")
    if env_device:
        return env_device

    try:
        import paddle

        if paddle.device.is_compiled_with_cuda():
            paddle.set_device("gpu:0")
            return "gpu:0"
    except Exception:
        pass
    return "cpu"


def _resolve_paddleocr_model_dirs() -> tuple[Path, Path]:
    env_det = os.environ.get("PADDLEOCR_DET_DIR")
    env_rec = os.environ.get("PADDLEOCR_REC_DIR")
    if env_det and env_rec:
        return Path(env_det), Path(env_rec)

    ascii_det = ASCII_MODEL_ROOT / "det" / "PP-OCRv5_server_det_infer"
    ascii_rec = ASCII_MODEL_ROOT / "rec" / "PP-OCRv5_server_rec_infer"
    if _is_valid_paddle_model_dir(ascii_det) and _is_valid_paddle_model_dir(ascii_rec):
        return ascii_det, ascii_rec

    project_det = PROJECT_MODEL_ROOT / "det" / "PP-OCRv5_server_det_infer"
    project_rec = PROJECT_MODEL_ROOT / "rec" / "PP-OCRv5_server_rec_infer"
    if _is_valid_paddle_model_dir(project_det) and _is_valid_paddle_model_dir(project_rec):
        return project_det, project_rec

    raise FileNotFoundError(
        "PaddleOCR PP-OCRv5 model dirs not found. Set PADDLEOCR_DET_DIR and PADDLEOCR_REC_DIR, "
        "or place models under models/paddleocr/{det,rec}."
    )


def _is_valid_paddle_model_dir(path: Path) -> bool:
    return (path / "inference.json").exists() and (path / "inference.pdiparams").exists() and (path / "inference.yml").exists()


def _run_paddleocr_image(path: Path, file_id: str) -> tuple[str, list[DocumentBlock], float]:
    engine = _paddleocr_engine()
    result = engine.predict(str(path))
    item = result[0] if isinstance(result, list) and result else result
    data = item if isinstance(item, dict) else item.to_dict()
    texts = [str(text) for text in (data.get("rec_texts") or [])]
    scores = [float(score) for score in (data.get("rec_scores") or [])]
    polys = data.get("rec_polys") or data.get("dt_polys") or []
    blocks: list[DocumentBlock] = []
    for index, text in enumerate(texts, start=1):
        confidence = scores[index - 1] if index - 1 < len(scores) else 0.0
        bbox = _poly_to_bbox(polys[index - 1]) if index - 1 < len(polys) else None
        blocks.append(
            DocumentBlock(
                file_id=file_id,
                block_id=f"{file_id}_OCR{index:04d}",
                page=1,
                block_type="image_text",
                text=text,
                bbox=bbox,
                source="paddleocr",
                confidence=round(confidence, 6),
                reading_order=index + 1,
            )
        )
    avg_confidence = round(sum(scores) / len(scores), 6) if scores else 0.0
    return "\n".join(texts), blocks, avg_confidence


def _poly_to_bbox(poly: object) -> list[float] | None:
    try:
        points = poly.tolist() if hasattr(poly, "tolist") else poly
        xs = [float(point[0]) for point in points]
        ys = [float(point[1]) for point in points]
        return [min(xs), min(ys), max(xs), max(ys)]
    except Exception:
        return None


def _read_txt(path: Path) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "utf-16", "latin-1"):
        try:
            return path.read_text(encoding=encoding), encoding
        except UnicodeError:
            continue
    return path.read_text(encoding="gb18030", errors="ignore"), "gb18030-ignore"


def _paragraph_blocks(file_id: str, paragraphs: list[str], source: str, start_order: int = 1) -> list[DocumentBlock]:
    blocks: list[DocumentBlock] = []
    order = start_order
    for paragraph in paragraphs:
        block_type = "title" if order == start_order and _is_heading(paragraph) else "paragraph"
        blocks.append(DocumentBlock(file_id, f"{file_id}_B{order:04d}", None, block_type, paragraph, None, source, 1.0, order))
        order += 1
    return blocks


def _split_paragraphs(text: str) -> list[str]:
    return [part.strip() for part in re.split(r"\n+", text) if part.strip()]


def _is_heading(text: str) -> bool:
    return len(text) <= 40 and bool(re.search(r"(通知|方案|报告|总结|申请|合同|办法|细则|说明|第[一二三四五六七八九十]+章)", text))


def _is_image_sidecar_text(text: str) -> bool:
    head = text[:500]
    return all(term in head for term in ("图片名称:", "图片编号:", "下载URL:"))


def _extract_title(file_name: str, paragraphs: list[str]) -> str:
    stem = Path(file_name).stem
    candidates = [stem, *(paragraph for paragraph in paragraphs[:5] if len(paragraph) <= 80)]
    for candidate in candidates:
        if any(term in candidate for term in ("通知", "方案", "报告", "总结", "申请", "合同", "办法", "细则", "预算", "会议")):
            return candidate[:120]
    return candidates[0][:120] if candidates else ""


def _summarize_table(table: list[list[str]]) -> str:
    headers = [cell for cell in (table[0] if table else []) if cell]
    joined = " ".join(" ".join(row) for row in table[:10])
    keywords = extract_keywords(joined, top_n=10)
    return f"表格字段:{'、'.join(headers[:20])}。主要关键词:{'、'.join(keywords)}。样例内容:{normalize_text(joined)[:500]}"


def _excel_sheet_text(sheet_name: str, n_rows: int, n_cols: int, headers: list[str], rows: list[tuple[object, ...]]) -> str:
    sample_values: list[str] = []
    for row in rows[:8]:
        for value in row[:8]:
            if value is not None:
                sample_values.append(str(value))
    sample = normalize_text(" ".join(sample_values))[:500]
    keywords = extract_keywords(sample, top_n=10)
    return (
        f"工作表名称:{sheet_name}。行数:{n_rows}。列数:{n_cols}。"
        f"表头字段:{'、'.join(headers[:20])}。主要关键词:{'、'.join(keywords)}。样例内容:{sample}"
    )


def _infer_headers(rows: list[tuple[object, ...]]) -> list[str]:
    for row in rows[:5]:
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if values:
            return values
    return []


def _docx_image_count(path: Path) -> int:
    try:
        with zipfile.ZipFile(path) as archive:
            return sum(1 for name in archive.namelist() if name.startswith("word/media/"))
    except Exception:
        return 0


def _pdf_page_count(path: Path) -> int:
    data = path.read_bytes()
    return max(1, len(re.findall(rb"/Type\s*/Page\b", data)))


def _extract_pdf_text_minimal(path: Path, max_chars: int) -> str:
    data = path.read_bytes()
    objects = _parse_pdf_objects(data)
    font_cmaps = _pdf_font_cmaps(objects)
    page_numbers = [number for number, body in sorted(objects.items()) if re.search(rb"/Type\s*/Page\b", body)]
    text_parts: list[str] = []
    for page_number in page_numbers:
        text_parts.append(_extract_pdf_page_text(objects, font_cmaps, page_number))
        if sum(len(part) for part in text_parts) >= max_chars:
            break
    return normalize_text("\n".join(text_parts))[:max_chars]


def _parse_pdf_objects(data: bytes) -> dict[int, bytes]:
    objects: dict[int, bytes] = {}
    for match in re.finditer(rb"(\d+)\s+(\d+)\s+obj\b(.*?)\bendobj", data, flags=re.S):
        objects[int(match.group(1))] = match.group(3)
    return objects


def _pdf_stream(body: bytes) -> bytes:
    if b"stream" not in body:
        return b""
    header, rest = body.split(b"stream", 1)
    raw = rest.split(b"endstream", 1)[0]
    if b"/FlateDecode" not in header:
        return raw.strip(b"\r\n")
    for candidate in (raw.strip(b"\r\n"), raw.lstrip(b"\r\n"), raw.lstrip(b"\r\n") + b"\r", raw.lstrip(b"\r\n") + b"\r\n"):
        try:
            return zlib.decompress(candidate)
        except Exception:
            continue
    return b""


def _pdf_font_cmaps(objects: dict[int, bytes]) -> dict[int, dict[str, str]]:
    cmaps: dict[int, dict[str, str]] = {}
    for number, body in objects.items():
        match = re.search(rb"/ToUnicode\s+(\d+)\s+0\s+R", body)
        if match and int(match.group(1)) in objects:
            cmaps[number] = _parse_pdf_cmap(_pdf_stream(objects[int(match.group(1))]))
    return cmaps


def _parse_pdf_cmap(stream: bytes) -> dict[str, str]:
    cmap: dict[str, str] = {}
    text = stream.decode("latin-1", errors="ignore")
    for block in re.finditer(r"beginbfchar(.*?)endbfchar", text, flags=re.S):
        for source, target in re.findall(r"<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>", block.group(1)):
            cmap[source.upper()] = _decode_pdf_unicode_hex(target)
    for block in re.finditer(r"beginbfrange(.*?)endbfrange", text, flags=re.S):
        for source, end, target in re.findall(r"<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>", block.group(1)):
            start_code = int(source, 16)
            end_code = int(end, 16)
            target_code = int(target, 16)
            width = len(source)
            for code in range(start_code, end_code + 1):
                cmap[f"{code:0{width}X}"] = _decode_pdf_unicode_hex(f"{target_code + code - start_code:04X}")
    return cmap


def _extract_pdf_page_text(objects: dict[int, bytes], font_cmaps: dict[int, dict[str, str]], page_number: int) -> str:
    page = objects[page_number]
    content_refs = _content_refs(page)
    font_refs = {name.decode("ascii", errors="ignore"): int(number) for name, number in re.findall(rb"/(FT\d+|F\d+)\s+(\d+)\s+0\s+R", page)}
    current_font: str | None = None
    chunks: list[str] = []
    token_pattern = re.compile(rb"/(FT\d+|F\d+)\s+[-0-9.]+\s+Tf|<([0-9A-Fa-f]+)>\s*Tj|\[(.*?)\]\s*TJ|\((.*?)\)\s*Tj|\bET\b", re.S)
    for content_ref in content_refs:
        content = _pdf_stream(objects.get(content_ref, b""))
        for match in token_pattern.finditer(content):
            if match.group(1):
                current_font = match.group(1).decode("ascii", errors="ignore")
            elif match.group(2):
                chunks.append(_decode_pdf_hex_text(match.group(2).decode("ascii"), font_cmaps.get(font_refs.get(current_font or "", -1), {})))
            elif match.group(3):
                cmap = font_cmaps.get(font_refs.get(current_font or "", -1), {})
                for hex_text in re.findall(rb"<([0-9A-Fa-f]+)>", match.group(3)):
                    chunks.append(_decode_pdf_hex_text(hex_text.decode("ascii"), cmap))
            elif match.group(4):
                chunks.append(match.group(4).decode("latin-1", errors="ignore"))
            else:
                chunks.append("\n")
    return "".join(chunks)


def _content_refs(page: bytes) -> list[int]:
    list_match = re.search(rb"/Contents\s*\[(.*?)\]", page, flags=re.S)
    if list_match:
        return [int(value) for value in re.findall(rb"(\d+)\s+0\s+R", list_match.group(1))]
    single_match = re.search(rb"/Contents\s+(\d+)\s+0\s+R", page)
    return [int(single_match.group(1))] if single_match else []


def _decode_pdf_hex_text(value: str, cmap: dict[str, str]) -> str:
    value = value.upper()
    output: list[str] = []
    i = 0
    while i < len(value):
        four = value[i : i + 4]
        two = value[i : i + 2]
        if four in cmap:
            output.append(cmap[four])
            i += 4
        elif two in cmap:
            output.append(cmap[two])
            i += 2
        else:
            try:
                output.append(bytes.fromhex(four).decode("utf-16-be"))
            except Exception:
                output.append("")
            i += 4
    return "".join(output)


def _decode_pdf_unicode_hex(value: str) -> str:
    try:
        return bytes.fromhex(value).decode("utf-16-be")
    except Exception:
        return chr(int(value, 16))


def _fullwidth_to_halfwidth(text: str) -> str:
    chars: list[str] = []
    for char in text:
        code = ord(char)
        if code == 0x3000:
            chars.append(" ")
        elif 0xFF01 <= code <= 0xFF5E:
            chars.append(chr(code - 0xFEE0))
        else:
            chars.append(char)
    return "".join(chars)


def _is_readable(path: Path) -> bool:
    try:
        with path.open("rb") as handle:
            handle.read(1)
        return True
    except OSError:
        return False


def _initial_type(extension: str) -> str:
    if extension in WORD_EXTENSIONS:
        return "word"
    if extension in EXCEL_EXTENSIONS:
        return "excel"
    if extension in PDF_EXTENSIONS:
        return "pdf"
    if extension in TEXT_EXTENSIONS:
        return "txt"
    if extension in IMAGE_EXTENSIONS:
        return "image"
    return "other"
